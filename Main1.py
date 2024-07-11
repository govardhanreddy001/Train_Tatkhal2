import openpyxl
workbook = openpyxl.load_workbook("Excel_Files/Copy of RAIL_WAY_RESERVATION.xlsx")
sheet = workbook['Sheet1']
total_rows = sheet.max_row
total_cols = sheet.max_column

print(total_rows)
print(total_cols)
print(sheet.cell(1,3).value)

for r in range(1,total_rows+1):
    for c in range(1,total_cols+1):
        print(sheet.cell(r,c).value,end= "    ")
    print(  )

